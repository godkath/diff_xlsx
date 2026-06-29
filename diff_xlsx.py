"""
diff_xlsx.py — 对比两个 xlsx/xls 第一个 sheet 的差异
- 自动找主键列 (优先 ZJHM 证件号码，可配置)
- 自动建 F1 <-> F2 的列名映射（不区分大小写 + 别名）
- 输出人员差异（主键集合）+ 全字段 cell 差异
- 数字差异智能识别：57.0 vs 57 视为相同（不报警）
- 支持 .xls (用 xlrd 1.2.0) 和 .xlsx (用 openpyxl)
- 可选 --out-xlsx 生成 Excel 差异表

用法:
    python diff_xlsx.py <文件1> <文件2> [标签1] [标签2]
    python diff_xlsx.py a.xlsx b.xlsx "中文版" "EN版" --out-xlsx diff.xlsx
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict

# 常见主键列名 (按优先级匹配, 优先 ZJHM 证件号码)
KEY_CANDIDATES = [
    "ZJHM", "证件号码", "身份证号", "身份证", "ID_NUMBER", "id_number",
    "sfzh", "SFZH", "ID", "id",
]

# 英文码 <-> 中文标签 别名映射（用于跨表字段匹配）
# 左边是 F1（库表）常见的英文码，右边是 F2（学信网）对应的中文标签
# 两侧都参与列名查找，顺序无关
COLUMN_ALIASES = {
    # 身份信息
    "KSBH":  ["考生编号", "考生号", "考号"],
    "XM":    ["姓名"],
    "ZJLX":  ["证件类型"],
    "ZJHM":  ["证件号码"],
    "CSRQ":  ["出生日期"],
    "XBM":   ["性别码"],
    "MZM":   ["民族码"],
    "ZZMMM": ["政治面貌码"],
    "HFM":   ["婚否码"],
    "XYJRM": ["现役军人码"],
    # 户籍档案
    "HKSZDM":   ["户口所在地码"],
    "HKSZDXXDZ":["户口所在地详细地址"],
    "DASZDW":   ["档案所在单位"],
    "DASZDWDZ": ["档案所在单位地址"],
    "DASZDWYZBM":["档案所在单位邮政编码"],
    # 毕业/学位
    "BYDWM": ["毕业学校代码"],
    "BYDW":  ["毕业学校名称"],
    "BYZYDM":["毕业专业代码"],
    "BYZYMC":["毕业专业名称"],
    "XXXS":  ["学习形式"],
    "XLM":   ["学历码"],
    "XLZSBH":["学历证书编号"],
    "BYNY":  ["毕业年月"],
    "ZCXH":  ["注册学号"],
    "XWM":   ["学位码"],
    "XWZSBH":["学位证书编号"],
    # 录取信息
    "LQDWDM": ["录取单位代码"],
    "LQDWMC": ["录取单位名称"],
    "LQZYDM": ["录取专业代码"],
    "LQZYMC": ["录取专业名称"],
    "LQYJFXDM":["录取研究方向码"],
    "LQYJFXMC":["录取研究方向名称"],
    "KSFSM":  ["招生类型"],
    "ZXJH":   ["专项计划"],
    "BLZGNX": ["保留入学资格年限"],
    "LQLBM":  ["录取类别码"],
    "DXWPDW": ["定向委培单位"],
    "DXWPDWSZDM":["定向委培单位所在地码"],
    "LQYXSM": ["录取院系所码"],
    "LHPYDW": ["联合培养单位"],
    "LHPYDWM":["联合培养单位码"],
    "FSCJ":   ["复试成绩"],
    "BZ":     ["备注"],
    "GGJHLB": ["专项计划2"],
}

# 反向映射：中文标签 -> 英文码
COLUMN_ALIASES_REV = {}
for en, zh_list in COLUMN_ALIASES.items():
    for zh in zh_list:
        COLUMN_ALIASES_REV[zh] = en


def find_col(headers, candidates):
    """在 headers 中找 candidates 之一，不区分大小写；返回索引或 None"""
    h_lookup = {(h or "").upper(): i for i, h in enumerate(headers)}
    for c in candidates:
        if c.upper() in h_lookup:
            return h_lookup[c.upper()]
    return None


def load_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    if not wb.sheetnames:
        wb.close()
        raise ValueError(f"{path} 没有 sheet")
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    key_idx = find_col(headers, KEY_CANDIDATES)
    if key_idx is None:
        wb.close()
        raise ValueError(
            f"{path} 找不到主键列（候选: {KEY_CANDIDATES}）\n实际 headers: {headers[:20]}"
        )
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        k = row[key_idx]
        if k is None:
            continue
        rows[str(k).strip()] = list(row)
    wb.close()
    return headers, rows, key_idx


def list_sheets(path):
    """列出文件中所有 sheet 名 (xls/xlsx 都支持)"""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xls':
        try:
            import xlrd
            wb = xlrd.open_workbook(path)
            return wb.sheet_names()
        except Exception as e:
            return [f"⚠️ 读取失败: {e}"]
    else:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            names = wb.sheetnames[:]
            wb.close()
            return names
        except Exception as e:
            return [f"⚠️ 读取失败: {e}"]


def load_any(path, sheet_name=None):
    """加载 xls/xlsx 文件指定 sheet, 返回 (headers, rows, key_idx)
    sheet_name 不传则取第一个 sheet。
    返回: rows 是 dict, key 为主键值 (如身份证号)
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xls':
        return _load_xls(path, sheet_name)
    else:
        # 复用 load_xlsx 逻辑, 但接受指定 sheet
        wb = load_workbook(path, read_only=True, data_only=True)
        if not wb.sheetnames:
            wb.close()
            raise ValueError(f"{path} 没有 sheet")
        sn = sheet_name or wb.sheetnames[0]
        if sn not in wb.sheetnames:
            wb.close()
            raise ValueError(f"{path} 没有 sheet '{sn}', 可用: {wb.sheetnames}")
        ws = wb[sn]
        headers = [c.value for c in ws[1]]
        key_idx = find_col(headers, KEY_CANDIDATES)
        if key_idx is None:
            wb.close()
            raise ValueError(
                f"{path} 找不到主键列（候选: {KEY_CANDIDATES}）\n实际 headers: {headers[:20]}"
            )
        rows = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            k = row[key_idx]
            if k is None:
                continue
            rows[str(k).strip()] = list(row)
        wb.close()
        return headers, rows, key_idx


def _load_xls(path, sheet_name=None):
    """读 .xls 文件 (用 xlrd 1.2.0)"""
    import xlrd
    wb = xlrd.open_workbook(path)
    if not wb.sheet_names():
        raise ValueError(f"{path} 没有 sheet")
    sn = sheet_name or wb.sheet_names()[0]
    if sn not in wb.sheet_names():
        raise ValueError(f"{path} 没有 sheet '{sn}', 可用: {wb.sheet_names()}")
    sh = wb.sheet_by_name(sn)
    if sh.nrows < 1:
        raise ValueError(f"{path}/{sn} 是空的")
    headers = list(sh.row_values(0))
    # 转列索引为 int (xlrd colx 是 int)
    key_idx = find_col(headers, KEY_CANDIDATES)
    if key_idx is None:
        raise ValueError(
            f"{path} 找不到主键列（候选: {KEY_CANDIDATES}）\n实际 headers: {headers[:20]}"
        )
    # xlrd 默认按 Excel 数值类型返回 (数字是 float, 日期是 float)
    # 我们统一转字符串供后面 cell() 函数处理
    rows = {}
    for r in range(1, sh.nrows):
        row = sh.row_values(r)
        k = row[key_idx]
        if k is None or k == '':
            continue
        # xlrd 数字转 int if integer
        if isinstance(k, float) and k.is_integer():
            k = str(int(k))
        else:
            k = str(k).strip()
        rows[k] = list(row)
    return headers, rows, key_idx


def cell(v):
    """统一 cell 值，数字 57.0 == 57 视为同值（read_only 模式下 cell 都是字符串，所以先尝试解析）"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    # 处理 "57.0" / "57.00" / "57.000" 这种情况
    if "." in s:
        try:
            f = float(s)
            if f.is_integer():
                return str(int(f))
        except ValueError:
            pass
    return s


def build_mapping(headers1, headers2):
    """
    建 headers1 -> headers2 的列名映射
    1. 优先按英文码↔中文别名（COLUMN_ALIASES）
    2. 其次按不区分大小写的精确匹配
    """
    h2_lookup = {(h or "").upper(): i for i, h in enumerate(headers2)}
    mapping = {}
    for i, h1 in enumerate(headers1):
        if h1 is None:
            continue
        key_upper = h1.upper()
        # 第一轮：精确大小写不敏感匹配
        if key_upper in h2_lookup:
            mapping[i] = h2_lookup[key_upper]
            continue
        # 第二轮：通过别名映射
        candidates = []
        if key_upper in {k.upper() for k in COLUMN_ALIASES}:
            # h1 是英文码，找其对应的中文别名
            for en, zh_list in COLUMN_ALIASES.items():
                if en.upper() == key_upper:
                    candidates.extend(zh_list)
        elif key_upper in {k.upper() for k in COLUMN_ALIASES_REV}:
            # h1 是中文标签，找其对应的英文码
            for zh, en in COLUMN_ALIASES_REV.items():
                if zh.upper() == key_upper:
                    candidates.append(en)
        found = None
        for c in candidates:
            cu = c.upper()
            if cu in h2_lookup:
                found = h2_lookup[cu]
                break
        mapping[i] = found
    return mapping


def generate_diff_xlsx(out_path: str, file1: str, sheet1: str, file2: str, sheet2: str,
                       label1: str, label2: str, headers1: list, rows1: dict, key_idx1: int,
                       headers2: list, rows2: dict, key_idx2: int) -> str:
    """生成 Excel 差异表 (4 个 sheet)
    Sheet 1: 人员差异
    Sheet 2: 字段结构差异
    Sheet 3: 字段值差异 (每行: 主键/字段/F1值/F2值)
    Sheet 4: 总结
    返回生成的路径
    """
    from openpyxl import Workbook
    from datetime import datetime

    wb = Workbook()
    # 默认 sheet 重命名为总结
    summary_ws = wb.active
    summary_ws.title = "总结"
    detail_ws = wb.create_sheet("人员差异")
    struct_ws = wb.create_sheet("字段结构差异")
    value_ws = wb.create_sheet("字段值差异")

    name_candidates = ["XM", "姓名", "NAME", "name"]
    n1 = find_col(headers1, name_candidates)
    n2 = find_col(headers2, name_candidates)

    s1, s2 = set(rows1), set(rows2)
    common = s1 & s2

    # ---- Sheet 1: 人员差异 ----
    detail_ws.append([f"仅 {label1} 有 (n={len(s1 - s2)})", f"仅 {label2} 有 (n={len(s2 - s1)})"])
    detail_ws.cell(1, 1).font = Font(bold=True, color="FFFFFF")
    detail_ws.cell(1, 1).fill = PatternFill("solid", fgColor="4472C4")
    detail_ws.cell(1, 2).font = Font(bold=True, color="FFFFFF")
    detail_ws.cell(1, 2).fill = PatternFill("solid", fgColor="4472C4")
    max_only = max(len(s1 - s2), len(s2 - s1), 1)
    only1 = sorted(s1 - s2)
    only2 = sorted(s2 - s1)
    for i in range(max_only):
        row = []
        if i < len(only1):
            k = only1[i]
            name = cell(rows1[k][n1]) if n1 is not None else ""
            row.append(f"{k}  {name}")
        else:
            row.append("")
        if i < len(only2):
            k = only2[i]
            name = cell(rows2[k][n2]) if n2 is not None else ""
            row.append(f"{k}  {name}")
        else:
            row.append("")
        detail_ws.append(row)

    # ---- Sheet 2: 字段结构差异 ----
    mapping = build_mapping(headers1, headers2)
    f1_only = [headers1[i] for i in sorted(mapping) if mapping[i] is None and headers1[i]]
    f2_only = [h for h in headers2
               if h and h.upper() not in {(c or "").upper() for c in headers1 if c}]
    struct_ws.append([f"{label1} 独有列 (n={len(f1_only)})", f"{label2} 独有列 (n={len(f2_only)})"])
    struct_ws.cell(1, 1).font = Font(bold=True, color="FFFFFF")
    struct_ws.cell(1, 1).fill = PatternFill("solid", fgColor="4472C4")
    struct_ws.cell(1, 2).font = Font(bold=True, color="FFFFFF")
    struct_ws.cell(1, 2).fill = PatternFill("solid", fgColor="4472C4")
    struct_max = max(len(f1_only), len(f2_only), 1)
    for i in range(struct_max):
        struct_ws.append([
            f1_only[i] if i < len(f1_only) else "",
            f2_only[i] if i < len(f2_only) else "",
        ])

    # ---- Sheet 3: 字段值差异 ----
    value_ws.append(["主键", "字段", f"{label1} 值", f"{label2} 值"])
    for c in range(1, 5):
        value_ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        value_ws.cell(1, c).fill = PatternFill("solid", fgColor="4472C4")

    if common:
        # 按 F1 列顺序遍历, 输出每个有差异的 (主键, 字段, F1 值, F2 值)
        for f1i, f2i in mapping.items():
            if f2i is None:
                continue
            f1c = headers1[f1i]
            for k in sorted(common):
                v1 = cell(rows1[k][f1i])
                v2 = cell(rows2[k][f2i])
                if v1 != v2:
                    value_ws.append([k, f1c, v1, v2])
    else:
        value_ws.append(["(无共有记录)", "", "", ""])

    # 列宽自适应
    for ws in [detail_ws, struct_ws, value_ws]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30
    value_ws.column_dimensions['A'].width = 25
    value_ws.column_dimensions['B'].width = 25

    # ---- Sheet 4: 总结 ----
    summary_ws.append(["项", "值"])
    summary_ws.cell(1, 1).font = Font(bold=True, color="FFFFFF")
    summary_ws.cell(1, 1).fill = PatternFill("solid", fgColor="4472C4")
    summary_ws.cell(1, 2).font = Font(bold=True, color="FFFFFF")
    summary_ws.cell(1, 2).fill = PatternFill("solid", fgColor="4472C4")
    summary_ws.append([f"生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_ws.append([f"文件 1", f"{file1} (sheet={sheet1})"])
    summary_ws.append([f"文件 2", f"{file2} (sheet={sheet2})"])
    summary_ws.append([f"标签 1", label1])
    summary_ws.append([f"标签 2", label2])
    summary_ws.append([f"文件 1 主键列", headers1[key_idx1] if key_idx1 is not None and key_idx1 < len(headers1) else ""])
    summary_ws.append([f"文件 2 主键列", headers2[key_idx2] if key_idx2 is not None and key_idx2 < len(headers2) else ""])
    summary_ws.append([f"文件 1 记录数", len(rows1)])
    summary_ws.append([f"文件 2 记录数", len(rows2)])
    summary_ws.append([f"人员差", abs(len(rows1) - len(rows2))])
    summary_ws.append([f"文件 1 列数", len(headers1)])
    summary_ws.append([f"文件 2 列数", len(headers2)])
    summary_ws.append([f"文件 1 独有列数", len(f1_only)])
    summary_ws.append([f"文件 2 独有列数", len(f2_only)])
    # 统计 cell 差异
    total_diff_cells = 0
    diff_field_count = 0
    for f1i, f2i in mapping.items():
        if f2i is None: continue
        cnt = sum(1 for k in common if cell(rows1[k][f1i]) != cell(rows2[k][f2i]))
        if cnt:
            diff_field_count += 1
            total_diff_cells += cnt
    summary_ws.append([f"有差异字段数", diff_field_count])
    summary_ws.append([f"总差异 cell 数", total_diff_cells])
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 50

    wb.save(out_path)
    return out_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description="对比两个 xls/xlsx 的差异")
    parser.add_argument("file1")
    parser.add_argument("file2")
    parser.add_argument("label1", nargs="?", default="F1")
    parser.add_argument("label2", nargs="?", default="F2")
    parser.add_argument("--sheet1", default=None, help="文件 1 sheet 名 (默认第一个)")
    parser.add_argument("--sheet2", default=None, help="文件 2 sheet 名 (默认第一个)")
    parser.add_argument("--out-xlsx", default=None, help="输出 Excel 差异表路径")
    args = parser.parse_args()

    p1, p2 = args.file1, args.file2
    l1, l2 = args.label1, args.label2

    print("=" * 70)
    print(f"=== {l1}  vs  {l2} ===")
    print("=" * 70)
    print(f"{l1}: {p1}")
    print(f"{l2}: {p2}\n")

    h1, r1, k1 = load_any(p1, args.sheet1)
    h2, r2, k2 = load_any(p2, args.sheet2)
    print(f"{l1} 列数={len(h1)} (主键 {h1[k1]} 在第 {k1} 列), 记录数={len(r1)}")
    print(f"{l2} 列数={len(h2)} (主键 {h2[k2]} 在第 {k2} 列), 记录数={len(r2)}\n")

    # 1. 人员差异
    s1, s2 = set(r1), set(r2)
    print("--- 人员差异（按主键集合）---")
    print(f"  共有: {len(s1 & s2)}")
    print(f"  仅 {l1} 有: {len(s1 - s2)}")
    print(f"  仅 {l2} 有: {len(s2 - s1)}")

    # 找姓名列（尝试常见列名）
    name_candidates = ["XM", "姓名", "NAME", "name"]
    n1 = find_col(h1, name_candidates)
    n2 = find_col(h2, name_candidates)

    if s1 - s2:
        print(f"\n  [仅 {l1} 有] (n={len(s1 - s2)}):")
        for k in sorted(s1 - s2):
            name = cell(r1[k][n1]) if n1 is not None else ""
            print(f"    {k}  {name}")
    if s2 - s1:
        print(f"\n  [仅 {l2} 有] (n={len(s2 - s1)}):")
        for k in sorted(s2 - s1):
            name = cell(r2[k][n2]) if n2 is not None else ""
            print(f"    {k}  {name}")

    # 2. 字段差异
    common = s1 & s2
    if not common:
        print("\n无共有记录，跳过字段比对")
        return

    mapping = build_mapping(h1, h2)
    f1_only = [h1[i] for i in sorted(mapping) if mapping[i] is None and h1[i]]
    f2_only = [
        h for h in h2
        if h and h.upper() not in {(c or "").upper() for c in h1 if c}
    ]
    print(f"\n--- 字段结构差异 ---")
    print(f"  {l1} 独有列 ({len(f1_only)}): {f1_only}")
    print(f"  {l2} 独有列 ({len(f2_only)}): {f2_only}")

    # 3. 全字段 cell 比对
    diff_count = Counter()
    diff_samples = defaultdict(list)

    for f1i, f2i in mapping.items():
        if f2i is None:
            continue
        f1c = h1[f1i]
        cnt = 0
        samples = []
        for key in common:
            v1 = cell(r1[key][f1i])
            v2 = cell(r2[key][f2i])
            if v1 != v2:
                cnt += 1
                if len(samples) < 3:
                    samples.append((key, v1, v2))
        if cnt > 0:
            diff_count[f1c] = cnt
            diff_samples[f1c] = samples

    print(f"\n--- 共有 {len(common)} 条记录的字段差异（按差异条数排序）---\n")
    if not diff_count:
        print("  ✅ 所有可比字段值完全一致")
    for k, v in diff_count.most_common():
        print(f"  {k}: {v} 条")
        for s in diff_samples[k]:
            print(f"    例: 主键={s[0]}  {l1}='{s[1]}'  {l2}='{s[2]}'")
        print()

    # 4. 总结
    print("=" * 70)
    print("【总结】")
    print(f"  人员: {l1}={len(r1)}  {l2}={len(r2)}  差 {abs(len(r1) - len(r2))} 人")
    print(f"  字段: {l1}={len(h1)} 列  {l2}={len(h2)} 列  {l1}独有 {len(f1_only)} {l2}独有 {len(f2_only)}")
    print(f"  数据: {len(diff_count)} 个字段有差异（总计 {sum(diff_count.values())} cell）")
    print("=" * 70)

    # 5. 可选：输出 Excel 差异表
    if args.out_xlsx:
        path = generate_diff_xlsx(
            args.out_xlsx, p1, args.sheet1 or "(第一个)", p2, args.sheet2 or "(第一个)",
            l1, l2, h1, r1, k1, h2, r2, k2
        )
        print(f"\n📊 Excel 差异表已生成: {path}")


if __name__ == "__main__":
    main()